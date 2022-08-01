import openpyxl
import re
import os
from pathlib import Path

os.chdir(r"C:\Users\Myscriptslocation")
wb = openpyxl.load_workbook('demo_wb.xlsx')
inventory = wb['Inventory']
fw = wb['FW']

# create a list of firewalls based on the first 5 chars and strip whitespaces
# In the for loop we can select the range of cells we want to include in the list

firewall_list = []
for row in fw['A2': f'A{fw.max_row}']:
    for cell in row:
        try:
            firewall_list.append(cell.value.strip()[0:7])
        except AttributeError:
            pass

firewall_keywords = re.compile("|".join(firewall_list),re.IGNORECASE)

# Get the value of a cell: inventory['c10'].value
# Get the total amount of rows with content: inventory.max_row
# inventory[f'd{i}'].value contains the values from the column "Device IP"

commands = []
for i in range(2,inventory.max_row,1):
    try:
        if "SDWAN" in inventory[f'b{i}'].value and \
        inventory[f'a{i}'].value == "Fortigate" and \
        firewall_keywords.findall(str(inventory[f'b{i}'].value)):
            print(f"{inventory[f'b{i}'].value} device ip is: {inventory[f'd{i}'].value}")
            commands.append(f"command: ssh admin@{str(inventory[f'd{i}'].value)} && zip -r /home/afa/algosec/firewalls/" + f"{str(inventory[f'b{i}'].value)[0:7]}") 
    except TypeError:
        pass
    
commands.sort()

with open(r'./commands.sh', 'w') as output_file:
    output_file.write(f'list of commands\n')
    [output_file.write(f'{cmdline}\n') for cmdline in commands]
    output_file.write('end of commands')
