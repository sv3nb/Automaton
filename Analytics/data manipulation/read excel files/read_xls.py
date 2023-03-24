# This script will read in an excel spreadsheet where each sheet is the location
# on each sheet we have 3 columns consisting of the vlan(id), vlan name and the IP subnet
# we then convert this into a python dictionary

import openpyxl
from pprint import pprint

wb = openpyxl.load_workbook('locations networks.xlsx')

networks = []
for sheet in wb.sheetnames: # wb.sheetnames  = [sheet1, sheet2, sheet3, ..]
    site = wb[sheet]
    for i in range(2,site.max_row):  # max.row counts the number of rows that are not empty
        vlan = {}
        vlan['vlanid'] = site[f'A{i}'].value
        vlan['vlan name'] = site[f'B{i}'].value
        vlan['subnet'] = site[f'C{i}'].value
        vlan['location'] = sheet
        networks.append(vlan)

[pprint(network) for network in networks]

'''

Here a sample output of the script

{'location': 'Antwerp',
 'subnet': '10.16.10.0/24',
 'vlan name': 'Office_users',
 'vlanid': 'VLAN10'}
{'location': 'Antwerp',
 'subnet': '10.16.20.0/24',
 'vlan name': 'Office_voice',
 'vlanid': 'VLAN20'}
{'location': 'Antwerp',
 'subnet': '10.16.30.128/25',
 'vlan name': 'Office_printers',
 'vlanid': 'VLAN30'}
{'location': 'Antwerp',
 'subnet': '10.16.40.40.0/24',
 'vlan name': 'Office_wireless',
 'vlanid': 'VLAN40'}
{'location': 'Antwerp',
 'subnet': '10.16.50.0/23',
 'vlan name': 'IOT_camera',
 'vlanid': 'VLAN50'}
{'location': 'Antwerp',
 'subnet': '10.16.65.0/24',
 'vlan name': 'IOT_packaging',
 'vlanid': 'VLAN65'}
{'location': 'Antwerp',
 'subnet': '172.16.70.0/24',
 'vlan name': 'DMZ_public',
 'vlanid': 'VLAN70'}
{'location': 'Antwerp',
 'subnet': '172.16.80.0/24',
 'vlan name': 'DMZ_internal',
 'vlanid': 'VLAN80'}
{'location': 'Antwerp',
 'subnet': '10.16.90.0/24',
 'vlan name': 'Office_remote',
 'vlanid': 'VLAN90'}
{'location': 'Brussels',
 'subnet': '10.16.10.0/24',
 'vlan name': 'Office_users',
 'vlanid': 'VLAN10'}
{'location': 'Brussels',
 'subnet': '10.16.20.0/24',
 'vlan name': 'Office_voice',
 'vlanid': 'VLAN20'}
'''
